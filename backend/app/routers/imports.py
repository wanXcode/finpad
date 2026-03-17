"""
Import router - Upload and parse CSV/Excel files
"""
import io
import csv
import json
import os
import tempfile
import subprocess
from datetime import datetime
from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException
from app.auth import get_current_user
from app.database import get_async_db

router = APIRouter(prefix="/api/import", tags=["import"])

PLATFORM_DETECT = {
    "alipay": ["交易号", "商家订单号", "交易创建时间"],
    "wechat": ["微信支付账单", "交易时间", "交易类型"],
    "cmb": ["招商银行", "记账日期", "交易金额"],
    "icbc": ["工商银行", "收入", "支出"],
}


def detect_platform(headers: list[str], content_preview: str) -> str:
    normalized_headers = [str(h).replace("\ufeff", "").strip() for h in headers]
    joined_headers = "|".join(normalized_headers)

    for h in normalized_headers:
        if "交易号" in h or "商家订单号" in h or "交易创建时间" in h:
            return "alipay"
        if "微信支付" in h or "交易单号" in h or "交易类型" in h:
            return "wechat"

    if any(k in joined_headers for k in ["交易号", "商家订单号", "交易创建时间", "支付宝"]):
        return "alipay"
    if any(k in joined_headers for k in ["微信支付", "交易单号", "交易类型"]):
        return "wechat"

    # Bank statement detection (generic: ICBC, CMB, etc.)
    bank_keywords = ["记账日期", "交易日期", "余额", "摘要", "交易摘要", "对方账户", "对方户名", "币种"]
    if sum(1 for k in bank_keywords if k in joined_headers) >= 2:
        if any(k in joined_headers or k in content_preview for k in ["工商银行", "工行"]):
            return "icbc"
        if any(k in joined_headers or k in content_preview for k in ["招商银行", "招行"]):
            return "cmb"
        return "bank"

    if any(k in content_preview for k in ["交易号", "商家订单号", "交易创建时间", "支付宝"]):
        return "alipay"
    if any(k in content_preview for k in ["微信支付", "交易单号", "交易类型"]):
        return "wechat"
    if any(k in content_preview for k in ["招商银行", "记账日期"]):
        return "cmb"
    if any(k in content_preview for k in ["工商银行"]):
        return "icbc"
    return "unknown"


def parse_csv_preview(content: bytes, filename: str):
    """Parse CSV and return preview data"""
    text = content.decode("utf-8-sig", errors="replace")
    lines = text.strip().split("\n")

    # Skip leading comment lines (Alipay CSVs have header comments)
    data_start = 0
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped and not stripped.startswith("#") and not stripped.startswith("-"):
            # Check if this looks like a header row
            if "," in stripped or "\t" in stripped:
                data_start = i
                break

    reader = csv.reader(lines[data_start:])
    rows = list(reader)
    if len(rows) < 2:
        raise ValueError("文件内容为空或格式不正确")

    headers = [h.strip() for h in rows[0]]
    data_rows = rows[1:]
    platform = detect_platform(headers, text[:500])

    return {
        "headers": headers,
        "preview_rows": [r for r in data_rows[:10]],
        "total_rows": len(data_rows),
        "platform": platform,
        "filename": filename,
    }


def parse_pdf_preview(content: bytes, filename: str):
    platform = "cmb" if "cmb" in filename.lower() or "招行" in filename else "icbc" if "icbc" in filename.lower() or "工行" in filename else "unknown"
    return {
        "headers": ["PDF账单预览"],
        "preview_rows": [["已识别为PDF文件，请确认平台后导入"]],
        "total_rows": 1,
        "platform": platform,
        "filename": filename,
    }


def parse_xlsx_preview(content: bytes, filename: str):
    """Parse Excel and return preview data"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c) if c is not None else "" for c in row])
        wb.close()
    except Exception as e:
        raise ValueError(f"无法解析 Excel 文件: {e}")

    if len(rows) < 2:
        raise ValueError("文件内容为空")

    # Find header row (skip comment rows)
    header_idx = 0
    for i, row in enumerate(rows):
        non_empty = [c for c in row if c.strip()]
        if len(non_empty) >= 3:
            header_idx = i
            break

    headers = [h.strip() for h in rows[header_idx]]
    data_rows = rows[header_idx + 1:]
    content_preview = str(rows[:5])
    platform = detect_platform(headers, content_preview)

    return {
        "headers": headers,
        "preview_rows": data_rows[:10],
        "total_rows": len(data_rows),
        "platform": platform,
        "filename": filename,
    }


@router.post("/upload")
async def upload_preview(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    content = await file.read()
    filename = file.filename or "unknown"

    if filename.lower().endswith(".csv"):
        preview = parse_csv_preview(content, filename)
    elif filename.lower().endswith((".xlsx", ".xls")):
        preview = parse_xlsx_preview(content, filename)
    elif filename.lower().endswith(".pdf"):
        preview = parse_pdf_preview(content, filename)
    else:
        raise HTTPException(400, "不支持的文件格式，请上传 CSV、Excel 或 PDF 文件")

    return preview


@router.post("/confirm")
async def confirm_import(
    file: UploadFile = File(...),
    platform: str = Form("alipay"),
    user: dict = Depends(get_current_user),
):
    if platform == "unknown":
        raise HTTPException(400, "未识别账单类型，请先手动选择正确的平台后再确认导入")

    content = await file.read()
    filename = file.filename or "unknown"
    db = await get_async_db()

    try:
        # Parse file
        if filename.lower().endswith(".csv"):
            text = content.decode("utf-8-sig", errors="replace")
            lines = text.strip().split("\n")
            data_start = 0
            for i, line in enumerate(lines):
                stripped = line.strip()
                if stripped and not stripped.startswith("#") and not stripped.startswith("-"):
                    if "," in stripped:
                        data_start = i
                        break
            reader = csv.reader(lines[data_start:])
            rows = list(reader)
            headers = [h.strip() for h in rows[0]]
            data_rows = rows[1:]
            total = len(data_rows)
            if platform == "alipay":
                created, skipped, failed = await _import_alipay(db, headers, data_rows, user["id"])
            elif platform == "wechat":
                created, skipped, failed = await _import_wechat(db, headers, data_rows, user["id"])
            elif platform in ("icbc", "cmb", "bank"):
                created, skipped, failed = await _import_bank_csv(db, headers, data_rows, platform, user["id"])
            else:
                raise HTTPException(400, f"暂不支持 {platform} 格式的CSV导入")
        elif filename.lower().endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            all_rows = []
            for row in ws.iter_rows(values_only=True):
                all_rows.append([str(c) if c is not None else "" for c in row])
            wb.close()
            header_idx = 0
            for i, row in enumerate(all_rows):
                non_empty = [c for c in row if c.strip()]
                if len(non_empty) >= 3:
                    header_idx = i
                    break
            headers = [h.strip() for h in all_rows[header_idx]]
            data_rows = all_rows[header_idx + 1:]
            total = len(data_rows)
            if platform == "wechat":
                created, skipped, failed = await _import_wechat(db, headers, data_rows, user["id"])
            elif platform in ("icbc", "cmb", "bank"):
                created, skipped, failed = await _import_bank_csv(db, headers, data_rows, platform, user["id"])
            else:
                raise HTTPException(400, f"暂不支持 {platform} 格式的Excel导入")
        elif filename.lower().endswith(".pdf"):
            total, created, skipped, failed = await _import_bank_pdf(db, content, filename, platform, user["id"])
        else:
            raise HTTPException(400, "不支持的文件格式")

        # Log import
        await db.execute("""
            INSERT INTO import_logs (user_id, filename, platform, file_size, total_records, created_records, skipped_records, failed_records, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'completed')
        """, (user["id"], filename, platform, len(content), total, created, skipped, failed))
        await db.commit()

        return {"total": total, "created": created, "skipped": skipped, "failed": failed}

    except HTTPException:
        raise
    except Exception as e:
        await db.execute("""
            INSERT INTO import_logs (user_id, filename, platform, file_size, status, error_message)
            VALUES (?, ?, ?, ?, 'error', ?)
        """, (user["id"], filename, platform, len(content), str(e)))
        await db.commit()
        raise HTTPException(500, f"导入失败: {e}")


async def _import_alipay(db, headers: list, rows: list, user_id: int):
    """Import Alipay CSV data"""
    # Build header index
    idx = {h: i for i, h in enumerate(headers)}
    created = skipped = failed = 0

    for row in rows:
        if not row or all(not c.strip() for c in row):
            continue
        try:
            tx_no = row[idx.get("交易号", 0)].strip() if "交易号" in idx else ""
            if not tx_no:
                failed += 1
                continue
            tx_id = f"alipay_{tx_no}"

            # Check duplicate within current user only
            cursor = await db.execute(
                "SELECT id FROM transactions WHERE user_id = ? AND tx_id = ?",
                (user_id, tx_id),
            )
            if await cursor.fetchone():
                skipped += 1
                continue

            tx_time = row[idx.get("交易创建时间", idx.get("交易时间", 1))].strip()
            amount_str = row[idx.get("金额（元）", idx.get("金额", 5))].strip().replace(",", "")
            amount = abs(float(amount_str)) if amount_str else 0

            direction_raw = row[idx.get("收/支", idx.get("资金状态", 6))].strip() if "收/支" in idx or "资金状态" in idx else ""
            if "支出" in direction_raw:
                direction = "支出"
            elif "收入" in direction_raw:
                direction = "收入"
            elif "不计" in direction_raw:
                direction = "不计收支"
            else:
                direction = "不计收支"

            category = row[idx.get("交易分类", 12)].strip() if "交易分类" in idx and idx["交易分类"] < len(row) else "其他"
            counterparty = row[idx.get("交易对方", 2)].strip() if "交易对方" in idx else ""
            note = row[idx.get("商品名称", idx.get("商品说明", 3))].strip() if "商品名称" in idx or "商品说明" in idx else ""

            await db.execute("""
                INSERT INTO transactions (user_id, tx_id, tx_time, platform, account, direction, amount, category, original_category, counterparty, note, source)
                VALUES (?, ?, ?, '支付宝', '余额宝', ?, ?, ?, ?, ?, ?, 'manual_upload')
            """, (user_id, tx_id, tx_time, direction, amount, category, category, counterparty, note))
            created += 1

        except Exception:
            failed += 1

    await db.commit()
    return created, skipped, failed


async def _import_bank_csv(db, headers: list, rows: list, platform: str, user_id: int):
    """Import bank CSV/Excel data (ICBC, CMB, generic bank)"""
    idx = {h.strip(): i for i, h in enumerate(headers)}
    platform_labels = {"icbc": "工商银行", "cmb": "招商银行", "bank": "银行"}
    platform_label = platform_labels.get(platform, "银行")
    created = skipped = failed = 0

    # Flexible column name matching
    def find_col(*candidates):
        for c in candidates:
            for h, i in idx.items():
                if c in h:
                    return i
        return None

    date_col = find_col("记账日期", "交易日期", "日期", "时间")
    amount_col = find_col("交易金额", "金额", "发生额")
    income_col = find_col("收入金额", "收入", "贷方金额", "贷方发生额")
    expense_col = find_col("支出金额", "支出", "借方金额", "借方发生额")
    balance_col = find_col("余额", "账户余额", "账面余额")
    summary_col = find_col("摘要", "交易摘要", "用途", "备注")
    counterparty_col = find_col("对方户名", "对方账户名称", "交易对方", "对方名称", "对方")
    currency_col = find_col("币种", "币别")

    if date_col is None:
        raise ValueError("未找到日期列，请确认账单格式")

    for row_idx, row in enumerate(rows):
        if not row or all(not str(c).strip() for c in row):
            continue
        try:
            date_val = str(row[date_col]).strip().replace("/", "-") if date_col is not None and date_col < len(row) else ""
            if not date_val or len(date_val) < 8:
                continue

            # Determine amount and direction
            amount = 0.0
            direction = "不计收支"
            if income_col is not None and expense_col is not None:
                inc_str = str(row[income_col]).strip().replace(",", "").replace("¥", "") if income_col < len(row) else ""
                exp_str = str(row[expense_col]).strip().replace(",", "").replace("¥", "") if expense_col < len(row) else ""
                inc = abs(float(inc_str)) if inc_str and inc_str not in ("", "-", "0", "0.00", "--") else 0
                exp = abs(float(exp_str)) if exp_str and exp_str not in ("", "-", "0", "0.00", "--") else 0
                if inc > 0:
                    amount, direction = inc, "收入"
                elif exp > 0:
                    amount, direction = exp, "支出"
            elif amount_col is not None and amount_col < len(row):
                amt_str = str(row[amount_col]).strip().replace(",", "").replace("¥", "")
                if amt_str and amt_str not in ("", "-", "--"):
                    val = float(amt_str)
                    amount = abs(val)
                    direction = "支出" if val < 0 else ("收入" if val > 0 else "不计收支")

            if amount == 0:
                failed += 1
                continue

            summary = str(row[summary_col]).strip() if summary_col is not None and summary_col < len(row) else ""
            counterparty = str(row[counterparty_col]).strip() if counterparty_col is not None and counterparty_col < len(row) else ""
            balance = str(row[balance_col]).strip() if balance_col is not None and balance_col < len(row) else ""

            import hashlib
            tx_hash = hashlib.sha1(f"{date_val}|{amount}|{direction}|{summary}|{counterparty}".encode()).hexdigest()[:12]
            tx_id = f"{platform}_{date_val.replace('-', '')}_{row_idx}_{tx_hash}"

            cursor = await db.execute(
                "SELECT id FROM transactions WHERE user_id = ? AND tx_id = ?",
                (user_id, tx_id),
            )
            if await cursor.fetchone():
                skipped += 1
                continue

            note_parts = [s for s in [summary, f"余额:{balance}" if balance else ""] if s]
            note = " | ".join(note_parts)

            await db.execute(
                """INSERT INTO transactions (user_id, tx_id, tx_time, platform, account, direction, amount, category, original_category, counterparty, note, source)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'manual_upload')""",
                (user_id, tx_id, date_val, platform_label, platform_label, direction, amount, "银行卡流水", "银行卡流水", counterparty, note),
            )
            created += 1
        except Exception:
            failed += 1

    await db.commit()
    return created, skipped, failed


async def _import_bank_pdf(db, content: bytes, filename: str, platform: str, user_id: int):
    if platform not in ("cmb", "icbc"):
        raise HTTPException(400, f"暂不支持 {platform} 格式的PDF导入")

    parser_name = "parse_cmb_pdf.js" if platform == "cmb" else "parse_icbc_pdf.js"
    platform_label = "招商银行" if platform == "cmb" else "工商银行"

    with tempfile.TemporaryDirectory() as tmpdir:
        pdf_path = os.path.join(tmpdir, filename)
        json_path = os.path.join(tmpdir, "out.json")
        with open(pdf_path, "wb") as f:
            f.write(content)

        cmd = ["node", f"/app/scripts/parse/{parser_name}", pdf_path, json_path]
        proc = subprocess.run(cmd, capture_output=True, text=True)
        if proc.returncode != 0:
            raise HTTPException(400, f"{platform_label}账单解析失败: {proc.stderr.strip() or proc.stdout.strip() or '未知错误'}")

        with open(json_path, "r", encoding="utf-8") as f:
            records = json.load(f)

    created = skipped = failed = 0
    total = len(records)
    for rec in records:
        try:
            tx_id = rec["tx_id"]
            cursor = await db.execute(
                "SELECT id FROM transactions WHERE user_id = ? AND tx_id = ?",
                (user_id, tx_id),
            )
            if await cursor.fetchone():
                skipped += 1
                continue

            await db.execute(
                """INSERT INTO transactions (user_id, tx_id, tx_time, platform, account, direction, amount, category, original_category, counterparty, note, source)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'manual_upload')""",
                (
                    user_id,
                    tx_id,
                    rec.get("tx_time_text", ""),
                    rec.get("platform", platform_label),
                    rec.get("account", platform_label),
                    rec.get("direction", "不计收支"),
                    float(rec.get("amount", 0) or 0),
                    rec.get("category", "银行卡流水"),
                    rec.get("category", "银行卡流水"),
                    rec.get("counterparty", ""),
                    rec.get("note", ""),
                ),
            )
            created += 1
        except Exception:
            failed += 1

    await db.commit()
    return total, created, skipped, failed


async def _import_wechat(db, headers: list, rows: list, user_id: int):
    """Import WeChat Excel/CSV data"""
    idx = {h: i for i, h in enumerate(headers)}
    created = skipped = failed = 0

    for row in rows:
        if not row or all(not str(c).strip() for c in row):
            continue
        try:
            tx_no = str(row[idx.get("交易单号", 8)]).strip() if "交易单号" in idx else ""
            if not tx_no or tx_no == "None":
                failed += 1
                continue
            tx_id = f"wechat_{tx_no}"

            cursor = await db.execute(
                "SELECT id FROM transactions WHERE user_id = ? AND tx_id = ?",
                (user_id, tx_id),
            )
            if await cursor.fetchone():
                skipped += 1
                continue

            tx_time = str(row[idx.get("交易时间", 0)]).strip()
            amount_str = str(row[idx.get("金额(元)", idx.get("金额", 5))]).strip().replace("¥", "").replace(",", "")
            amount = abs(float(amount_str)) if amount_str else 0

            direction_raw = str(row[idx.get("收/支", 4)]).strip() if "收/支" in idx else ""
            if "支出" in direction_raw:
                direction = "支出"
            elif "收入" in direction_raw:
                direction = "收入"
            else:
                direction = "不计收支"

            category = str(row[idx.get("交易类型", 1)]).strip() if "交易类型" in idx else "其他"
            counterparty = str(row[idx.get("交易对方", 2)]).strip() if "交易对方" in idx else ""
            note = str(row[idx.get("商品", 3)]).strip() if "商品" in idx else ""

            await db.execute("""
                INSERT INTO transactions (user_id, tx_id, tx_time, platform, account, direction, amount, category, original_category, counterparty, note, source)
                VALUES (?, ?, ?, '微信', '零钱', ?, ?, ?, ?, ?, ?, 'manual_upload')
            """, (user_id, tx_id, tx_time, direction, amount, category, category, counterparty, note))
            created += 1

        except Exception:
            failed += 1

    await db.commit()
    return created, skipped, failed


@router.get("/history")
async def import_history(
    user: dict = Depends(get_current_user),
):
    db = await get_async_db()
    cursor = await db.execute("""
        SELECT id, filename, platform, total_records, created_records, skipped_records, status, created_at
        FROM import_logs WHERE user_id = ? ORDER BY created_at DESC LIMIT 20
    """, (user["id"],))
    rows = await cursor.fetchall()
    items = []
    for r in rows:
        items.append({
            "id": r[0], "filename": r[1], "platform": r[2],
            "total_records": r[3], "created_records": r[4],
            "skipped_records": r[5], "status": r[6], "created_at": r[7],
        })
    return {"items": items}
